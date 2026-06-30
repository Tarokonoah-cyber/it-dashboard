from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\User\Desktop\通訊錄.xlsx")

DEPARTMENT_CODES = {
    "EO": "EO",
    "GRO": "GRO",
    "SMD": "SMD",
    "SPA": "SPA",
    "人資": "人資",
    "人力資源": "人資",
    "HR": "人資",
    "司機": "司機",
    "娛樂部": "娛樂部",
    "安全室": "安全室",
    "客務部": "GRO",
    "Front Office": "GRO",
    "工程部": "工程部",
    "Engineering": "工程部",
    "廚房": "廚房",
    "Kitchen": "廚房",
    "房務部": "房務部",
    "Housekeeping": "房務部",
    "採購部": "採購部",
    "Purchasing": "採購部",
    "行李員": "行李員",
    "Bell": "行李員",
    "訂房": "訂房",
    "Reservation": "訂房",
    "財務": "財務部",
    "Finance": "財務部",
}


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def has_email(text):
    return "@" in text


def looks_like_phone(text):
    compact = re.sub(r"\s+", "", text)
    return bool(re.search(r"\d{3,}", compact))


def department_code(text):
    for needle, code in DEPARTMENT_CODES.items():
        if needle in text:
            return code
    return ""


def split_title(text):
    text = clean(text)
    if not text:
        return "", ""
    match = re.match(r"^(.+?)\s{2,}(.+)$", text)
    if match:
        return clean(match.group(1)), clean(match.group(2))
    return text, ""


def row_values(sheet, row_index):
    return [clean(sheet.cell(row_index, col).value) for col in range(1, 10)]


def parse_contacts(workbook_path: Path):
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    contacts = []
    current_department = ""
    current_group = ""

    for row_index in range(1, sheet.max_row + 1):
        title_zh, title_en, name_zh, name_en, extension, office_phone, cht_mobile, mobile_phone, email = row_values(
            sheet, row_index
        )
        row_text = " ".join(
            part
            for part in [title_zh, title_en, name_zh, name_en, extension, office_phone, cht_mobile, mobile_phone, email]
            if part
        )
        if not row_text or "Extension" in row_text and "E-mail" in row_text:
            continue

        detected_department = department_code(row_text)
        has_contact_data = any([name_zh, name_en, extension, office_phone, cht_mobile, mobile_phone, email])

        if detected_department and not name_zh and not extension and not office_phone and not cht_mobile and not mobile_phone:
            current_department = detected_department
            current_group = title_zh
            if has_email(title_en) and not email:
                email = title_en
                title_en = ""
            if email:
                contacts.append(
                    {
                        "record_key": f"contacts-{row_index:04d}",
                        "department": current_department,
                        "title_zh": title_zh,
                        "title_en": title_en,
                        "name_zh": "",
                        "name_en": "",
                        "extension": "",
                        "office_phone": "",
                        "cht_mobile": "",
                        "mobile_phone": "",
                        "email": email,
                        "note": "",
                    }
                )
            continue

        if not current_department and detected_department:
            current_department = detected_department

        if not has_contact_data:
            current_group = title_zh or current_group
            continue

        if has_email(title_en) and not email:
            email = title_en
            title_en = ""

        if has_email(title_zh) and not email:
            email = title_zh
            title_zh = current_group or title_zh

        if not current_department:
            current_department = detected_department or "未分類"

        if not title_en:
            parsed_title_zh, parsed_title_en = split_title(title_zh)
            title_zh = parsed_title_zh
            title_en = parsed_title_en

        if title_zh and not name_zh and not extension and not office_phone and not cht_mobile and not mobile_phone and not email:
            current_group = title_zh
            continue

        contacts.append(
            {
                "record_key": f"contacts-{row_index:04d}",
                "department": current_department,
                "title_zh": title_zh,
                "title_en": title_en,
                "name_zh": name_zh,
                "name_en": name_en,
                "extension": extension,
                "office_phone": office_phone,
                "cht_mobile": cht_mobile,
                "mobile_phone": mobile_phone,
                "email": email,
                "note": "",
            }
        )

    return contacts


def extension_rows(rows):
    return [
        {
            "record_key": row["record_key"],
            "department": row["department"],
            "title_zh": row["title_zh"],
            "title_en": row["title_en"],
            "name_zh": row["name_zh"],
            "name_en": row["name_en"],
            "email": row["email"],
            "extension": row["extension"],
        }
        for row in rows
        if row.get("extension")
    ]


def load_env(path: Path):
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.lstrip().startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def supabase_request(url, service_key, table, query="", method="GET", body=None):
    endpoint = f"{url.rstrip('/')}/rest/v1/{table}"
    if query:
        endpoint = f"{endpoint}?{query}"
    request = urllib.request.Request(endpoint, method=method)
    request.add_header("apikey", service_key)
    request.add_header("Authorization", f"Bearer {service_key}")
    request.add_header("Content-Type", "application/json")
    request.add_header("Prefer", "return=representation,resolution=merge-duplicates")
    if body is not None:
        request.data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    with urllib.request.urlopen(request, timeout=30) as response:
        text = response.read().decode("utf-8")
        return json.loads(text) if text else []


def import_contacts(rows):
    project_dir = Path(__file__).resolve().parents[1]
    load_env(project_dir / ".env.local")
    supabase_url = os.environ.get("SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in vercel-dashboard/.env.local")

    batch_size = 100
    imported = 0
    for start in range(0, len(rows), batch_size):
        batch = rows[start : start + batch_size]
        supabase_request(
            supabase_url,
            service_key,
            "contacts",
            "on_conflict=record_key",
            method="POST",
            body=batch,
        )
        imported += len(batch)
        time.sleep(0.1)
    return imported


def main():
    parser = argparse.ArgumentParser(description="Import company contacts from 通訊錄.xlsx into Supabase contacts.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--json", type=Path, help="Write parsed rows to a JSON file.")
    parser.add_argument("--extensions-json", type=Path, help="Write only extension overlay rows to a JSON file.")
    parser.add_argument("--import", dest="do_import", action="store_true", help="Import parsed rows to Supabase.")
    args = parser.parse_args()

    rows = parse_contacts(args.workbook)
    print(f"parsed={len(rows)}")
    by_department = {}
    for row in rows:
        by_department[row["department"]] = by_department.get(row["department"], 0) + 1
    print("departments=" + json.dumps(by_department, ensure_ascii=False, sort_keys=True))
    print(json.dumps(rows[:8], ensure_ascii=False, indent=2))

    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"wrote={args.json}")

    if args.extensions_json:
        extensions = extension_rows(rows)
        args.extensions_json.parent.mkdir(parents=True, exist_ok=True)
        args.extensions_json.write_text(json.dumps(extensions, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"extensions={len(extensions)}")
        print(f"wrote={args.extensions_json}")

    if args.do_import:
        try:
            imported = import_contacts(rows)
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")
            print(f"import_error={error.code} {detail}", file=sys.stderr)
            raise
        print(f"imported={imported}")


if __name__ == "__main__":
    main()
